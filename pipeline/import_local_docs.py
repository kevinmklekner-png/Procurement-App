"""
Import local SOW, solicitation, and forecast PDF/DOCX files into the database.

Reuses the existing doc_parser.py pipeline for text extraction, classification,
and structured analysis.  Forecast PDFs get special tabular parsing.

Usage:
    python3 import_local_docs.py                  # Import all local files
    python3 import_local_docs.py --dry-run        # Preview without DB writes
    python3 import_local_docs.py --files f1 f2    # Import specific files
"""

import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Optional

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'analysis'))
from database import ProcurementDatabase
from doc_parser import (
    extract_text,
    classify_document,
    build_sow_analysis,
    extract_evaluation_factors,
)
import json

APP_DIR = Path(__file__).resolve().parent.parent

# Files known to be invalid (HTML error pages masquerading as PDFs)
INVALID_FILES = {
    "sec_forecast.pdf",
    "sec_forecast2.pdf",
    "sec_forecast_final.pdf",
    "sec_procurement_opps.pdf",
    "cfpb_forecast.pdf",
    "cfpb_forecast_2023.pdf",
    "sec_forecast3.xlsx",
}

# Forecast files that need special tabular parsing
FORECAST_FILES = {
    "contracting-forecast.pdf",
}

# Agency inference: filename patterns -> agency name
AGENCY_PATTERNS = [
    (r"^FHF[-\s]", "FHFA"),
    (r"FHFA", "FHFA"),
    (r"\bSEC\b", "SEC"),
    (r"\bOCC\b", "OCC"),
    (r"\bCFPB\b", "CFPB"),
    (r"\bFCA\b", "FCA"),
    # OCC solicitation number patterns (2031JW anywhere in filename)
    (r"2031JW", "OCC"),
    # SEC solicitation number prefix (503102...)
    (r"50310[12]", "SEC"),
]

# Text-based agency detection keywords
AGENCY_TEXT_SIGNALS = {
    "FHFA": ["federal housing finance agency", "fhfa"],
    "SEC": ["securities and exchange commission", "sec.gov", " the sec's ", " sec regional office", "50310225r", "50310224r", "50310226r"],
    "OCC": ["office of the comptroller of the currency", "comptroller of the currency", "occ.gov"],
    "CFPB": ["consumer financial protection bureau", "cfpb"],
    "FCA": ["farm credit administration", "fca.gov"],
}

# SEC forecast value category mapping
VALUE_CATEGORIES = {
    "A": (0, 250_000),
    "B": (250_000, 1_000_000),
    "C": (1_000_000, 2_500_000),
    "D": (2_500_000, 5_000_000),
    "E": (5_000_000, None),
}


def get_solicitation_id(db, notice_id: str):
    """Look up the solicitations table id for a notice_id."""
    cursor = db.conn.cursor()
    cursor.execute("SELECT id FROM solicitations WHERE notice_id = ?", (notice_id,))
    row = cursor.fetchone()
    return row["id"] if row else None


def store_document(db, doc: dict):
    """Insert a document record and return its id."""
    cursor = db.conn.cursor()
    try:
        cursor.execute('''
            INSERT OR IGNORE INTO opportunity_documents
                (solicitation_id, notice_id, filename, file_url, file_type,
                 doc_role, raw_text, description_html,
                 download_status, parse_status, error_message,
                 created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            doc.get("solicitation_id"),
            doc["notice_id"],
            doc.get("filename", ""),
            doc["file_url"],
            doc.get("file_type", ""),
            doc.get("doc_role", "unknown"),
            doc.get("raw_text", ""),
            doc.get("description_html", ""),
            doc.get("download_status", "complete"),
            doc.get("parse_status", "complete"),
            doc.get("error_message", ""),
            datetime.now().isoformat(),
            datetime.now().isoformat(),
        ))
        db.conn.commit()
        return cursor.lastrowid if cursor.lastrowid else None
    except Exception as e:
        print(f"  Error storing document: {e}")
        db.conn.rollback()
        return None


def store_sow_analysis(db, analysis: dict):
    """Insert a SOW analysis record."""
    cursor = db.conn.cursor()
    cursor.execute('''
        INSERT INTO sow_analysis
            (document_id, solicitation_id, notice_id,
             scope_summary, period_of_performance, place_of_performance,
             key_tasks, labor_categories, deliverables,
             compliance_reqs, ordering_mechanism, billing_instructions,
             confidence_score, extraction_method, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        analysis["document_id"], analysis["solicitation_id"], analysis["notice_id"],
        analysis["scope_summary"], analysis["period_of_performance"],
        analysis["place_of_performance"],
        analysis["key_tasks"], analysis["labor_categories"], analysis["deliverables"],
        analysis["compliance_reqs"], analysis["ordering_mechanism"],
        analysis["billing_instructions"],
        analysis["confidence_score"], analysis["extraction_method"],
        analysis["created_at"],
    ))
    db.conn.commit()


def store_eval_factors(db, factors):
    """Insert evaluation criteria records."""
    cursor = db.conn.cursor()
    for f in factors:
        cursor.execute('''
            INSERT INTO evaluation_criteria
                (document_id, solicitation_id, notice_id,
                 evaluation_phase, factor_number, factor_name, factor_weight,
                 subfactors, description, page_limit, rating_method, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            f["document_id"], f["solicitation_id"], f["notice_id"],
            f["evaluation_phase"], f["factor_number"], f["factor_name"],
            f["factor_weight"], f["subfactors"], f["description"],
            f["page_limit"], f["rating_method"], f["created_at"],
        ))
    db.conn.commit()


def run_analysis(db, doc_id, sol_id, notice_id, role, text):
    """Run appropriate analysis based on document role."""
    if role in ("sow", "pws"):
        analysis = build_sow_analysis(doc_id, sol_id, notice_id, text)
        store_sow_analysis(db, analysis)

    if role in ("solicitation", "evaluation_criteria"):
        factors = extract_evaluation_factors(doc_id, sol_id, notice_id, text)
        if factors:
            store_eval_factors(db, factors)
            print(f"    Extracted {len(factors)} evaluation factors")

    # For solicitations, also try SOW extraction (they often contain both)
    if role == "solicitation":
        analysis = build_sow_analysis(doc_id, sol_id, notice_id, text)
        if json.loads(analysis["labor_categories"]) or json.loads(analysis["key_tasks"]):
            store_sow_analysis(db, analysis)
            print(f"    Also extracted SOW data from solicitation")


def find_local_files() -> List[Path]:
    """Find all PDF, DOCX, and XLSX files in the app directory."""
    files = []
    for ext in ("*.pdf", "*.docx", "*.xlsx", "*.PDF", "*.DOCX", "*.XLSX"):
        files.extend(APP_DIR.glob(ext))
    return sorted(set(files))


def is_valid_pdf(file_bytes: bytes) -> bool:
    """Check if file bytes start with a PDF header (not HTML)."""
    return file_bytes[:5] == b"%PDF-"


def infer_agency(filename: str, text: str) -> str:
    """Infer agency from filename patterns and text content."""
    # Try filename patterns first
    for pattern, agency in AGENCY_PATTERNS:
        if re.search(pattern, filename, re.IGNORECASE):
            return agency

    # Fall back to text content analysis
    lower_text = text[:5000].lower()
    for agency, signals in AGENCY_TEXT_SIGNALS.items():
        for signal in signals:
            if signal in lower_text:
                return agency

    return "UNKNOWN"


def generate_notice_id(filename: str, agency: str) -> str:
    """Generate a synthetic notice ID for a local file."""
    # Strip extension and clean up URL-encoded characters
    slug = Path(filename).stem
    slug = slug.replace("+", "-").replace(" ", "-")
    # Remove redundant separators
    slug = re.sub(r"-{2,}", "-", slug).strip("-")
    # Truncate to reasonable length
    if len(slug) > 60:
        slug = slug[:60].rstrip("-")
    return f"LOCAL-{agency}-{slug}"


def parse_sec_forecast(text: str) -> List[dict]:
    """Parse SEC contracting forecast PDF into structured rows.

    Each data line looks like:
      CF Training Coordinator Support Services B FEDERAL MANAGEMENT PARTNERS, LLC 50310219C0033 1

    Key identifying features of data lines:
    - Contains a contract number matching 50310XXXXXXXXX
    - Ends with a quarter digit (1-4)
    - Contains a value category letter (A-E) before the incumbent name
    """
    rows = []

    # Pattern: Office Description ValueCat Incumbent ContractNum Quarter
    # Contract numbers are 50310 followed by 8+ chars
    line_pattern = re.compile(
        r'^(\w{2,5})\s+'           # Office abbreviation
        r'(.+?)\s+'                # Description (greedy but stops at value cat)
        r'([A-E])\s+'             # Value category
        r'(.+?)\s+'               # Incumbent
        r'(50310\w{8,})\s+'       # Contract number
        r'([1-4])\s*$'            # Quarter
    )

    for line in text.split("\n"):
        line = line.strip()
        if not line or '50310' not in line:
            continue

        m = line_pattern.match(line)
        if not m:
            continue

        office = m.group(1)
        description = m.group(2).strip()
        value_cat = m.group(3)
        incumbent = m.group(4).strip()
        contract_num = m.group(5)
        quarter = f"Q{m.group(6)}"

        low, high = VALUE_CATEGORIES.get(value_cat, (None, None))

        rows.append({
            "agency": "SEC",
            "office_name": office,
            "project_description": description,
            "estimated_amount_category": value_cat,
            "estimated_value_low": low,
            "estimated_value_high": high,
            "estimated_quarter": quarter,
            "fiscal_year": 2024,
            "source_document": "contracting-forecast.pdf (local import)",
            "notes": f"Incumbent: {incumbent}; Contract: {contract_num}",
        })

    return rows


# ---------------------------------------------------------------------------
# Excel pricing sheet parsing
# ---------------------------------------------------------------------------

# Known solicitation links: filename substring → notice_id
EXCEL_SOLICITATION_MAP = {
    "Attachment+D": "LOCAL-FHFA-Attachment-D-Pricing-Worksheet",
    "Attachment+C": "LOCAL-FHFA-Attachment-C-Revised-Price-Quotation-Worksheet",
    "Appian": "LOCAL-OCC-RFP-Attachment-2-Appian-Platform-and-Product-Support-IDIQ",
}


def _infer_agency_from_excel(wb, filename: str) -> str:
    """Infer agency from Excel workbook content or filename."""
    # Check solicitation map first (Attachment C/D → FHFA)
    for substring, nid in EXCEL_SOLICITATION_MAP.items():
        if substring in filename:
            if "FHFA" in nid:
                return "FHFA"
            if "OCC" in nid:
                return "OCC"
            if "SEC" in nid:
                return "SEC"

    # Filename-based
    upper = filename.upper()
    if "OCC" in upper or "APPIAN" in upper:
        return "OCC"
    if "FHF" in upper or "FHFA" in upper:
        return "FHFA"
    if "SEC" in upper:
        return "SEC"

    # Content-based: scan first sheet header rows
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        text = " ".join(str(c) for c in row if c)
        if "OCC" in text or "Comptroller of the Currency" in text:
            return "OCC"
        if "FHFA" in text or "Federal Housing Finance" in text:
            return "FHFA"
        if "SEC" in text or "Securities and Exchange" in text:
            return "SEC"
        if "CFPB" in text:
            return "CFPB"
        if "FCA" in text or "Farm Credit" in text:
            return "FCA"

    return "UNKNOWN"


def _detect_header_columns(row_values):
    """Detect which columns contain category, rate, hours, price, CLIN from a header row.

    Returns dict with keys: category, rate, hours, price, clin, title,
    site, catalog_id (column indices).
    """
    cols = {}
    for i, val in enumerate(row_values):
        if val is None:
            continue
        v = str(val).lower().strip()
        if "labor category" in v and "title" not in v:
            cols["category"] = i
        elif "clin" in v:
            cols["clin"] = i
        elif "vendor" in v and "title" in v:
            cols["title"] = i
        elif any(kw in v for kw in ("hourly rate", "unit price", "labor rate")):
            cols["rate"] = i
        elif any(kw in v for kw in ("estimated hours", "quantity", "total hours")):
            cols["hours"] = i
        elif any(kw in v for kw in ("extended", "total price", "extended total")):
            cols["price"] = i
        # GSA price list columns
        elif v == "title" and "category" not in cols:
            cols["category"] = i
        elif "gsa price with iff" in v and "rate" not in cols:
            cols["rate"] = i
        elif "predominant work site" in v:
            cols["site"] = i
        elif "unique catalog item id" in v:
            cols["catalog_id"] = i
    return cols


def _is_data_stop(val):
    """Check if a cell value indicates the end of labor category rows."""
    if val is None:
        return True
    s = str(val).strip().lower()
    return s == "" or any(kw in s for kw in ("total", "subtotal", "section ii", "section iii",
                                               "travel", "other direct", "grand total",
                                               "reimbursable"))


def _parse_period_name(sheet_name: str) -> tuple:
    """Extract a period name and number from a sheet name.

    Returns (period_name, period_number).
    """
    sn = sheet_name.strip()

    # "Ordering Period I" -> ("Ordering Period I", 1)
    import re as _re
    m = _re.search(r"Ordering\s+Period\s+([IVX]+|\d+)", sn, _re.IGNORECASE)
    if m:
        raw = m.group(1)
        roman = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5}
        num = roman.get(raw.upper(), int(raw) if raw.isdigit() else 0)
        return (sn, num)

    # "Year VI" / "Year VII" etc.
    m = _re.search(r"Year\s+([IVX]+|\d+)", sn, _re.IGNORECASE)
    if m:
        raw = m.group(1)
        roman = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5,
                 "VI": 6, "VII": 7, "VIII": 8, "IX": 9, "X": 10}
        num = roman.get(raw.upper(), int(raw) if raw.isdigit() else 0)
        return (sn, num)

    # "1st Year of Performance" etc.
    m = _re.search(r"(\d+)\w*\s+Year", sn, _re.IGNORECASE)
    if m:
        num = int(m.group(1))
        return (sn, num)

    return (sn, 0)


def _detect_site_type(sheet_name: str):
    """Detect site type from sheet name."""
    sn = sheet_name.lower()
    if "contractor" in sn:
        return "contractor"
    if "government" in sn:
        return "government"
    return None


def parse_pricing_sheet(ws, sheet_name: str) -> List[dict]:
    """Parse a single worksheet and return labor category rows.

    Auto-detects format by scanning the first ~10 rows for header patterns.
    Returns list of dicts with keys: category_name, category_title, clin_number,
    hourly_rate, estimated_hours, extended_price, period_name, period_number, site_type.
    """
    rows_data = []
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # Find header row (scan first 10 rows)
    header_cols = {}
    header_row_idx = None
    for i, row in enumerate(all_rows[:10]):
        cols = _detect_header_columns(row)
        if "category" in cols or ("clin" in cols and len(cols) >= 2):
            header_cols = cols
            header_row_idx = i
            break

    # Pattern A (Attachment C): no explicit "Labor Category" header column.
    # Column 0 has category name, column 1 has unit price, col 2 has quantity, col 3 extended total
    if not header_cols:
        # Check if row 1 has "Unit Price" in col 1 and row 2 has "Labor Categories" section header
        for i, row in enumerate(all_rows[:5]):
            text = " ".join(str(c) for c in row if c).lower()
            if "unit price" in text:
                header_cols = {"category": 0, "rate": 1, "hours": 2, "price": 3}
                header_row_idx = i
                break

    if not header_cols:
        return []

    # Find where data starts (skip section headers)
    data_start = header_row_idx + 1
    for i in range(header_row_idx + 1, min(header_row_idx + 5, len(all_rows))):
        row = all_rows[i]
        cat_idx = header_cols.get("category", 0)
        val = row[cat_idx] if cat_idx < len(row) else None
        if val is not None:
            s = str(val).strip().lower()
            if "section" in s or "labor categor" in s or s == "":
                data_start = i + 1
                continue
            break

    period_name, period_number = _parse_period_name(sheet_name)
    site_type = _detect_site_type(sheet_name)

    cat_idx = header_cols.get("category", 0)
    rate_idx = header_cols.get("rate")
    hours_idx = header_cols.get("hours")
    price_idx = header_cols.get("price")
    clin_idx = header_cols.get("clin")
    title_idx = header_cols.get("title")
    site_idx = header_cols.get("site")

    # For OCC pattern: rate columns may be per-year (Year 1-5)
    # Detect multi-year rate columns
    year_rate_cols = []
    if header_row_idx is not None:
        hrow = all_rows[header_row_idx]
        for ci, val in enumerate(hrow):
            if not val:
                continue
            vs = str(val).lower()
            if "year" in vs and "rate" in vs:
                # Extract year number
                m = re.search(r"year\s+(\d+)", vs)
                yr = int(m.group(1)) if m else 0
                year_rate_cols.append((ci, yr))

        # GSA price list: "GSA Price with IFF (A)", "(B)", etc. map to years
        # Also check the row below for year labels like "Year 6", "Year 7"
        gsa_rate_cols = []
        gsa_current_col = None
        for ci, val in enumerate(hrow):
            if val and "gsa price with iff" in str(val).lower():
                letter_m = re.search(r"\(([A-Z])\)", str(val))
                if letter_m:
                    gsa_rate_cols.append(ci)
                elif "as of" in str(val).lower() or "generation" in str(val).lower():
                    gsa_current_col = ci
        if gsa_rate_cols and header_row_idx + 1 < len(all_rows):
            year_row = all_rows[header_row_idx + 1]
            for ci in gsa_rate_cols:
                if ci < len(year_row) and year_row[ci]:
                    m = re.search(r"year\s+(\d+)", str(year_row[ci]).lower())
                    yr = int(m.group(1)) if m else 0
                    year_rate_cols.append((ci, yr))
            # Include the "current rate" column as a year entry
            if gsa_current_col is not None:
                # Determine current year from the year_row entries
                existing_years = [yr for _, yr in year_rate_cols if yr > 0]
                # Current rate is typically the active year (max of detected years or nearby)
                current_yr = max(existing_years) - len(existing_years) if existing_years else 0
                if current_yr <= 0:
                    current_yr = min(existing_years) if existing_years else 0
                year_rate_cols.insert(0, (gsa_current_col, current_yr))
            # Advance data start past the year label rows
            if year_rate_cols:
                data_start = max(data_start, header_row_idx + 3)

    # GSA-style dense tables: skip empty rows instead of stopping
    has_gsa_cols = bool(year_rate_cols and any("catalog_id" in header_cols for _ in [1]))

    for i in range(data_start, len(all_rows)):
        row = all_rows[i]
        if cat_idx >= len(row):
            continue
        cat_name = row[cat_idx]
        if _is_data_stop(cat_name):
            # For OCC pattern, empty category with a CLIN means blank template row — skip
            if clin_idx is not None and clin_idx < len(row) and row[clin_idx]:
                continue
            # For GSA pricelists / large tables, skip empty rows rather than stopping
            if "catalog_id" in header_cols:
                continue
            break

        cat_name = str(cat_name).strip()
        if not cat_name or "labor categor" in cat_name.lower() or cat_name.lower().startswith("example"):
            continue

        def _num(idx):
            if idx is None or idx >= len(row) or row[idx] is None:
                return None
            try:
                return float(row[idx])
            except (ValueError, TypeError):
                return None

        # Per-row site type (GSA price list has "Contractor_Facility" etc.)
        row_site = site_type
        if site_idx is not None and site_idx < len(row) and row[site_idx]:
            sv = str(row[site_idx]).lower()
            if "contractor" in sv:
                row_site = "contractor"
            elif "customer" in sv or "government" in sv:
                row_site = "government"
            elif "virtual" in sv:
                row_site = "virtual"
            else:
                row_site = str(row[site_idx]).strip()

        # CLIN or catalog ID
        clin_val = ""
        if clin_idx is not None and clin_idx < len(row) and row[clin_idx]:
            clin_val = str(row[clin_idx]).strip()
        elif "catalog_id" in header_cols:
            ci = header_cols["catalog_id"]
            if ci < len(row) and row[ci]:
                clin_val = str(row[ci]).strip()

        # Multi-year rate columns (OCC / GSA pattern)
        if year_rate_cols:
            for col_idx, yr_num in year_rate_cols:
                entry = {
                    "category_name": cat_name,
                    "category_title": str(row[title_idx]).strip() if title_idx and title_idx < len(row) and row[title_idx] else "",
                    "clin_number": clin_val,
                    "hourly_rate": _num(col_idx),
                    "estimated_hours": _num(hours_idx),
                    "extended_price": _num(price_idx),
                    "period_name": f"Year {yr_num}",
                    "period_number": yr_num,
                    "site_type": row_site,
                }
                rows_data.append(entry)
        else:
            entry = {
                "category_name": cat_name,
                "category_title": str(row[title_idx]).strip() if title_idx and title_idx < len(row) and row[title_idx] else "",
                "clin_number": clin_val,
                "hourly_rate": _num(rate_idx),
                "estimated_hours": _num(hours_idx),
                "extended_price": _num(price_idx),
                "period_name": period_name,
                "period_number": period_number,
                "site_type": row_site,
            }
            rows_data.append(entry)

    return rows_data


def import_excel_pricing(db: ProcurementDatabase, filepath: Path, dry_run: bool) -> int:
    """Import an Excel pricing worksheet into the labor_categories table."""
    import openpyxl

    filename = filepath.name
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    agency = _infer_agency_from_excel(wb, filename)
    notice_id = generate_notice_id(filename, agency)

    # Try to link to existing solicitation
    sol_id = None
    for substring, nid in EXCEL_SOLICITATION_MAP.items():
        if substring in filename:
            sol_id = get_solicitation_id(db, nid)
            if sol_id:
                notice_id = nid
            break

    # Skip summary/total/reference sheets
    skip_sheets = {"overall summary", "total evaluated contract price", "pricing reference",
                   "contractor info"}

    all_rows = []
    sheet_names = wb.sheetnames
    is_gsa = "pricelist" in filename.lower() or "price list" in " ".join(sheet_names).lower()
    for sn in sheet_names:
        if sn.strip().lower() in skip_sheets:
            continue
        ws = wb[sn]
        rows = parse_pricing_sheet(ws, sn)
        all_rows.extend(rows)

    wb.close()

    print(f"  Agency: {agency} | Notice: {notice_id} | Sheets: {len(sheet_names)}")
    print(f"  Parsed {len(all_rows)} labor category rows")

    if dry_run:
        seen = set()
        for r in all_rows:
            key = r["category_name"]
            if key not in seen:
                rate_str = f"${r['hourly_rate']:.2f}/hr" if r['hourly_rate'] else "template"
                hours_str = f"{r['estimated_hours']:.0f}h" if r['estimated_hours'] else ""
                clin_str = f"CLIN {r['clin_number']}" if r['clin_number'] else ""
                parts = [p for p in [clin_str, rate_str, hours_str] if p]
                print(f"    - {key} ({', '.join(parts)}) [{r['period_name']}]")
                seen.add(key)
        if len(seen) < len(all_rows):
            print(f"    ... {len(all_rows) - len(seen)} additional period rows")
        return len(all_rows)

    # Ensure solicitation record exists
    if not sol_id:
        sol_data = {
            "notice_id": notice_id,
            "solicitation_number": Path(filename).stem.replace("+", " "),
            "title": Path(filename).stem.replace("+", " "),
            "department": agency,
            "type_of_notice": "pricing_worksheet",
            "data_source": "local_import",
            "url": f"local://{filename}",
        }
        sol_id = db.insert_solicitation(sol_data)
        if not sol_id:
            sol_id = get_solicitation_id(db, notice_id)

    # Detect if this is a GSA price list (has "pricelist" in name or vendor info sheet)
    is_gsa = "pricelist" in filename.lower() or "price list" in " ".join(wb.sheetnames).lower()

    inserted = 0
    for row in all_rows:
        row["notice_id"] = notice_id
        row["solicitation_id"] = sol_id
        row["source_file"] = filename
        row["agency"] = agency
        row["data_source"] = "gsa_pricelist" if is_gsa else "excel_import"
        row_id = db.insert_labor_category(row)
        if row_id:
            inserted += 1

    print(f"  Inserted {inserted} rows ({len(all_rows) - inserted} duplicates skipped)")
    return inserted


def import_forecast(db: ProcurementDatabase, filepath: Path, dry_run: bool) -> int:
    """Import a forecast PDF into the forecast_opportunities table."""
    filename = filepath.name
    file_bytes = filepath.read_bytes()

    if not is_valid_pdf(file_bytes):
        print(f"  SKIP (invalid PDF): {filename}")
        return 0

    text = extract_text(filename, file_bytes)
    if not text.strip():
        print(f"  SKIP (no text extracted): {filename}")
        return 0

    rows = parse_sec_forecast(text)
    print(f"  Parsed {len(rows)} forecast entries from {filename}")

    if dry_run:
        for r in rows[:5]:
            print(f"    - {r['project_description'][:60]}... [{r['estimated_amount_category']}]")
        if len(rows) > 5:
            print(f"    ... and {len(rows) - 5} more")
        return len(rows)

    inserted = 0
    for row in rows:
        row_id = db.insert_forecast_opportunity(row)
        if row_id:
            inserted += 1

    print(f"  Inserted {inserted} new entries ({len(rows) - inserted} duplicates skipped)")
    return inserted


def import_solicitation(db: ProcurementDatabase, filepath: Path, dry_run: bool) -> bool:
    """Import a SOW/solicitation/PWS document into the database."""
    filename = filepath.name
    file_bytes = filepath.read_bytes()

    # Validate PDFs
    if filename.lower().endswith(".pdf") and not is_valid_pdf(file_bytes):
        print(f"  SKIP (invalid PDF): {filename}")
        return False

    # Extract text
    text = extract_text(filename, file_bytes)
    if not text or len(text.strip()) < 50:
        print(f"  SKIP (no/minimal text): {filename}")
        return False

    # Classify and infer agency
    doc_role = classify_document(filename, text)
    agency = infer_agency(filename, text)

    # If role is unknown but filename suggests RFQ/RFP, classify as solicitation
    if doc_role == "unknown":
        lower_name = filename.lower()
        if any(kw in lower_name for kw in ("rfq", "rfp", "solicitation")):
            doc_role = "solicitation"

    notice_id = generate_notice_id(filename, agency)
    file_url = f"local://{filename}"

    # Determine file type
    ext = filepath.suffix.lower().lstrip(".")
    file_type = ext if ext in ("pdf", "docx") else "unknown"

    print(f"  Role: {doc_role} | Agency: {agency} | ID: {notice_id}")
    print(f"  Text: {len(text)} chars")

    if dry_run:
        return True

    # 1. Insert solicitation record
    sol_data = {
        "notice_id": notice_id,
        "solicitation_number": Path(filename).stem.replace("+", " "),
        "title": Path(filename).stem.replace("+", " "),
        "description": text[:2000],
        "department": agency,
        "type_of_notice": doc_role,
        "data_source": "local_import",
        "url": file_url,
    }
    sol_row_id = db.insert_solicitation(sol_data)
    if not sol_row_id:
        # May already exist (INSERT OR REPLACE) — look it up
        sol_row_id = get_solicitation_id(db, notice_id)

    # 2. Store document record
    doc_id = store_document(db, {
        "solicitation_id": sol_row_id,
        "notice_id": notice_id,
        "filename": filename,
        "file_url": file_url,
        "file_type": file_type,
        "doc_role": doc_role,
        "raw_text": text,
    })

    if not doc_id:
        # Document may already exist (INSERT OR IGNORE on file_url)
        cursor = db.conn.cursor()
        cursor.execute("SELECT id FROM opportunity_documents WHERE file_url = ?", (file_url,))
        row = cursor.fetchone()
        if row:
            print(f"  Already imported (skipping analysis)")
            return True
        print(f"  WARNING: Failed to store document record")
        return False

    # 3. Run analysis pipeline (same as collect_documents.py)
    run_analysis(db, doc_id, sol_row_id, notice_id, doc_role, text)
    print(f"  Imported successfully")
    return True


def main():
    parser = argparse.ArgumentParser(
        description="Import local SOW, solicitation, and forecast files into the database"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview what would be imported without writing to DB")
    parser.add_argument("--files", nargs="+",
                        help="Import specific files (by filename, not full path)")
    args = parser.parse_args()

    # Find files to process
    all_files = find_local_files()

    if args.files:
        target_names = set(args.files)
        files = [f for f in all_files if f.name in target_names]
        missing = target_names - {f.name for f in files}
        if missing:
            print(f"WARNING: Files not found: {', '.join(missing)}")
    else:
        files = all_files

    if not files:
        print("No PDF/DOCX files found to import.")
        return

    print(f"{'[DRY RUN] ' if args.dry_run else ''}Found {len(files)} file(s) to process\n")

    db = ProcurementDatabase()

    stats = {"imported": 0, "skipped": 0, "forecast": 0, "excel": 0, "errors": 0}

    for filepath in files:
        filename = filepath.name
        print(f"\n--- {filename} ---")

        try:
            # Skip temp/lock files
            if filename.startswith("~$"):
                print(f"  SKIP (temp/lock file)")
                stats["skipped"] += 1
                continue

            # Skip known invalid files
            if filename in INVALID_FILES:
                print(f"  SKIP (known invalid / HTML error page)")
                stats["skipped"] += 1
                continue

            # Forecast files get special handling
            if filename in FORECAST_FILES:
                count = import_forecast(db, filepath, args.dry_run)
                stats["forecast"] += count
                continue

            # Excel pricing worksheets
            if filename.lower().endswith(".xlsx"):
                count = import_excel_pricing(db, filepath, args.dry_run)
                stats["excel"] += count
                continue

            # Everything else: SOW / solicitation / PWS import
            success = import_solicitation(db, filepath, args.dry_run)
            if success:
                stats["imported"] += 1
            else:
                stats["skipped"] += 1

        except Exception as e:
            print(f"  ERROR: {e}")
            stats["errors"] += 1

    db.close()

    print(f"\n{'=' * 50}")
    print(f"{'[DRY RUN] ' if args.dry_run else ''}Import Summary:")
    print(f"  Solicitations/SOWs imported: {stats['imported']}")
    print(f"  Excel labor category rows: {stats['excel']}")
    print(f"  Forecast entries: {stats['forecast']}")
    print(f"  Skipped: {stats['skipped']}")
    print(f"  Errors: {stats['errors']}")


if __name__ == "__main__":
    main()
